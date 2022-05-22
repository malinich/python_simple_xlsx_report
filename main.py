import uvicorn
import xlsxtopy
import uuid

from fastapi.params import Body
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

import settings
from fastapi import FastAPI, APIRouter, Depends, Request
from fastapi.responses import FileResponse

from db import get_session

SQL = """
    select c2.country_id, c2.country, sum(p.amount) from customer cus
    join payment p on cus.customer_id = p.customer_id
    join address a on a.address_id = cus.address_id
    join city c on c.city_id = a.city_id
    join country c2 on c2.country_id = c.country_id
    group by c2.country_id, c2.country
    order by sum(p.amount) DESC;
"""

SQL_2 = """
SELECT * from payment
    LEFT JOIN customer cus on cus.customer_id = payment.customer_id
    LEFT JOIN address a on cus.address_id = a.address_id
    LEFT JOIN city cit on a.city_id = cit.city_id
    LEFT JOIN country con on cit.country_id = con.country_id
WHERE con.country_id = :country_id
ORDER BY payment.payment_date DESC ;
"""


def ru_vary():
    rrows = []
    for _, row in enumerate(range(10000)):
        rrows.append([str(v) for v in range(5)])

    f_name = xlsxtopy.write_xlsx(rrows)
    return f_name


def py_vary():
    wb = Workbook()
    sheet = wb.active
    for idx, row in enumerate(range(10000)):
        for idx_r, rrow in enumerate(range(5)):
            sheet.cell(idx + 1, idx_r + 1, rrow)
    f_name = str(uuid.uuid4()) + ".xlsx"
    wb.save(f_name)
    return f_name


app = FastAPI(title="test for xlsx")

router = APIRouter()


@router.post("/report-rust", description="report for coutries")
async def report_counties(
        db_session=Depends(get_session),
        data: dict = Body(None),
        request: Request = None
):
    rows = await db_session.execute(SQL_2, {"country_id": data.get("country_id", -1)})
    f_name = xlsxtopy.write_xlsx([[str(v) for v in d] for d in rows])

    return FileResponse(f_name, media_type="application/vnd.openxmlformats")


@router.post("/report-python", description="report for coutries")
async def report_counties(
        db_session: AsyncSession = Depends(get_session),
        data: dict = Body({"country_id": None}),
        request: Request = None
):
    rows = await db_session.execute(SQL_2, {"country_id": data.get("country_id", -1)})
    wb = Workbook(write_only=True)
    sheet = wb.create_sheet()

    for row in rows:
        sheet.append(list(row))
    f_name = str(uuid.uuid4()) + ".xlsx"
    wb.save(f_name)
    return FileResponse(f_name, media_type="application/vnd.openxmlformats")


app.include_router(router)

if __name__ == "__main__":
    uvicorn.run("main:app",
                loop="uvloop",
                host=settings.SERVER_HOST,
                port=settings.SERVER_PORT,
                reload=True)
